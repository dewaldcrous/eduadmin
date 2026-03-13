import csv
import io
import json
from datetime import datetime

from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Q
from rest_framework import generics, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser

from django.core.exceptions import ObjectDoesNotExist

from .models import LessonPlan, LessonDelivery, PlanApprovalLog, LessonPlanAttachment
from .serializers import (
    LessonPlanListSerializer, LessonPlanDetailSerializer,
    SubmitPlanSerializer, ApprovePlanSerializer, DeliverLessonSerializer,
    LessonPlanAttachmentSerializer,
)
from apps.accounts.permissions import IsTeacherOrAbove, IsHODOrAbove


# ─── CRUD ────────────────────────────────────────────────────────────────────

class LessonPlanListCreateView(generics.ListCreateAPIView):
    permission_classes = [IsTeacherOrAbove]

    def get_serializer_class(self):
        if self.request.method == "POST":
            return LessonPlanDetailSerializer
        return LessonPlanListSerializer

    def get_queryset(self):
        user = self.request.user
        qs = LessonPlan.objects.select_related(
            "timetable_slot", "timetable_slot__teacher",
            "timetable_slot__subject", "timetable_slot__classroom",
            "timetable_slot__classroom__grade",
        ).order_by("-created_at")

        if user.role == "teacher":
            qs = qs.filter(timetable_slot__teacher=user)
        elif user.role == "hod":
            qs = qs.filter(timetable_slot__subject__hod=user)
        elif user.role == "grade_head":
            qs = qs.filter(timetable_slot__classroom__grade__grade_head=user)
        elif user.role in ["deputy", "principal", "admin"]:
            qs = qs.filter(timetable_slot__timetable__school=user.school)

        slot_id = self.request.query_params.get("slot")
        if slot_id:
            qs = qs.filter(timetable_slot_id=slot_id)

        plan_status = self.request.query_params.get("status")
        if plan_status:
            qs = qs.filter(status=plan_status)

        return qs


class LessonPlanDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsTeacherOrAbove]
    serializer_class = LessonPlanDetailSerializer

    def get_queryset(self):
        return LessonPlan.objects.select_related(
            "timetable_slot", "timetable_slot__teacher",
            "timetable_slot__subject", "timetable_slot__classroom",
            "timetable_slot__classroom__grade", "approved_by",
        ).prefetch_related("approval_logs", "curriculum_outcomes", "attachments")


# ─── WEEKLY VIEW ─────────────────────────────────────────────────────────────

class WeeklyPlanView(APIView):
    permission_classes = [IsTeacherOrAbove]

    def get(self, request):
        user = request.user
        teacher_id = request.query_params.get("teacher", user.id)

        if str(teacher_id) != str(user.id) and user.role == "teacher":
            teacher_id = user.id

        from apps.schools.models import TimetableSlot, Timetable, TimetableConfig

        timetable = Timetable.objects.filter(
            school=user.school, status="active"
        ).order_by("-year", "-term").first()

        if not timetable:
            return Response({"error": "No active timetable found"}, status=404)

        # Get timetable configuration for dynamic days/periods
        try:
            config = timetable.config
            cycle_type = config.cycle_type
            num_days = config.num_days
            periods_per_day = config.periods_per_day
        except ObjectDoesNotExist:
            # Default to weekly schedule
            cycle_type = "week"
            num_days = 5
            periods_per_day = 7

        # Build dynamic days list based on cycle type
        if cycle_type == "cycle":
            days = [f"D{i}" for i in range(1, num_days + 1)]
            day_labels = {f"D{i}": f"Day {i}" for i in range(1, num_days + 1)}
        else:
            week_days = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]
            week_labels = {
                "MON": "Monday", "TUE": "Tuesday", "WED": "Wednesday",
                "THU": "Thursday", "FRI": "Friday", "SAT": "Saturday", "SUN": "Sunday",
            }
            days = week_days[:num_days]
            day_labels = {d: week_labels[d] for d in days}

        slots = TimetableSlot.objects.filter(
            timetable=timetable, teacher_id=teacher_id
        ).select_related("subject", "classroom", "classroom__grade").order_by("period")

        plans = LessonPlan.objects.filter(
            timetable_slot__in=slots
        ).select_related(
            "timetable_slot", "timetable_slot__subject",
            "timetable_slot__classroom", "timetable_slot__classroom__grade",
        ).prefetch_related("attachments")

        plan_map = {}
        for plan in plans:
            key = f"{plan.timetable_slot.day}-{plan.timetable_slot.period}"
            if key not in plan_map or plan.version > plan_map[key].version:
                plan_map[key] = plan

        # Build periods list from actual slot data or defaults
        all_periods = set()
        period_times = {}
        for slot in slots:
            all_periods.add(slot.period)
            period_times[slot.period] = {
                "start": slot.start_time.strftime("%H:%M"),
                "end": slot.end_time.strftime("%H:%M"),
            }

        # If no slots, use default periods
        if not all_periods:
            all_periods = set(range(1, periods_per_day + 1))

        periods_list = sorted(all_periods)

        weekly = []
        for day in days:
            day_slots = slots.filter(day=day).order_by("period")
            day_data = {"day": day, "label": day_labels.get(day, day), "periods": []}
            for slot in day_slots:
                key = f"{day}-{slot.period}"
                plan = plan_map.get(key)
                period_data = {
                    "slot_id": slot.id,
                    "period": slot.period,
                    "start_time": slot.start_time.strftime("%H:%M"),
                    "end_time": slot.end_time.strftime("%H:%M"),
                    "subject": slot.subject.name,
                    "subject_code": slot.subject.code,
                    "classroom": str(slot.classroom),
                    "is_break": slot.is_break,
                    "plan": None,
                }
                if plan:
                    period_data["plan"] = {
                        "id": plan.id,
                        "title": plan.title,
                        "status": plan.status,
                        "objectives": plan.objectives,
                        "activities": plan.activities,
                        "differentiation": plan.differentiation,
                        "resources_note": plan.resources_note,
                        "plan_type": plan.plan_type,
                        "version": plan.version,
                        "has_delivery": hasattr(plan, "delivery"),
                        "has_reflection": hasattr(plan, "reflection"),
                        "attachment_count": plan.attachments.count(),
                    }
                day_data["periods"].append(period_data)
            weekly.append(day_data)

        total_slots = slots.filter(is_break=False).count()
        plans_done = len([p for p in plan_map.values() if p.status in ["approved", "draft", "pending"]])
        plans_approved = len([p for p in plan_map.values() if p.status == "approved"])

        return Response({
            "teacher_id": int(teacher_id),
            "timetable_id": timetable.id,
            "term": timetable.term,
            "year": timetable.year,
            "config": {
                "cycle_type": cycle_type,
                "num_days": num_days,
                "periods_per_day": periods_per_day,
                "days": [{"key": d, "label": day_labels.get(d, d)} for d in days],
                "periods": [{"period": p, **period_times.get(p, {"start": "", "end": ""})} for p in periods_list],
            },
            "weekly": weekly,
            "stats": {
                "total_slots": total_slots,
                "plans_created": plans_done,
                "plans_approved": plans_approved,
                "plans_missing": total_slots - plans_done,
                "completion_rate": round((plans_done / total_slots) * 100) if total_slots > 0 else 0,
            },
        })


# ─── APPROVAL WORKFLOW ───────────────────────────────────────────────────────

class SubmitPlanView(APIView):
    permission_classes = [IsTeacherOrAbove]

    def post(self, request):
        serializer = SubmitPlanSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            plan = LessonPlan.objects.get(id=serializer.validated_data["plan_id"])
        except LessonPlan.DoesNotExist:
            return Response({"error": "Plan not found"}, status=404)

        if plan.status not in ["draft", "rejected"]:
            return Response({"error": f"Cannot submit plan with status '{plan.status}'"}, status=400)

        plan.submit_for_approval()
        PlanApprovalLog.objects.create(lesson_plan=plan, action="submitted", actioned_by=request.user)
        return Response({"status": "submitted", "plan_id": plan.id})


class ApprovePlanView(APIView):
    permission_classes = [IsHODOrAbove]

    def post(self, request):
        serializer = ApprovePlanSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            plan = LessonPlan.objects.get(id=serializer.validated_data["plan_id"])
        except LessonPlan.DoesNotExist:
            return Response({"error": "Plan not found"}, status=404)

        if plan.status != "pending":
            return Response({"error": f"Cannot approve plan with status '{plan.status}'"}, status=400)

        action = serializer.validated_data["action"]
        feedback = serializer.validated_data.get("feedback", "")

        if action == "approved":
            plan.approve(request.user)
        elif action in ["rejected", "returned"]:
            plan.reject(request.user, feedback)

        PlanApprovalLog.objects.create(
            lesson_plan=plan, action=action, actioned_by=request.user, feedback=feedback,
        )
        return Response({"status": action, "plan_id": plan.id})


class DeliverLessonView(APIView):
    permission_classes = [IsTeacherOrAbove]

    def post(self, request):
        serializer = DeliverLessonSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            plan = LessonPlan.objects.get(id=serializer.validated_data["plan_id"])
        except LessonPlan.DoesNotExist:
            return Response({"error": "Plan not found"}, status=404)

        delivery, _ = LessonDelivery.objects.update_or_create(
            lesson_plan=plan,
            defaults={
                "completion": serializer.validated_data["completion"],
                "coverage_percent": serializer.validated_data["coverage_percent"],
                "delivered_by": request.user,
            },
        )

        if delivery.completion in ["partial", "not_done"]:
            try:
                from apps.reflections.services import create_carry_over
                create_carry_over(delivery, None)
            except Exception:
                pass

        return Response({
            "status": "delivered", "plan_id": plan.id,
            "completion": delivery.completion,
        })


class PendingApprovalsView(generics.ListAPIView):
    permission_classes = [IsHODOrAbove]
    serializer_class = LessonPlanListSerializer

    def get_queryset(self):
        user = self.request.user
        qs = LessonPlan.objects.filter(status="pending").select_related(
            "timetable_slot", "timetable_slot__teacher",
            "timetable_slot__subject", "timetable_slot__classroom",
            "timetable_slot__classroom__grade",
        ).order_by("created_at")

        if user.role == "hod":
            qs = qs.filter(timetable_slot__subject__hod=user)
        elif user.role == "grade_head":
            qs = qs.filter(timetable_slot__classroom__grade__grade_head=user)
        elif user.role in ["deputy", "principal", "admin"]:
            qs = qs.filter(timetable_slot__timetable__school=user.school)
        return qs


# ─── FILE ATTACHMENTS ────────────────────────────────────────────────────────

class AttachmentUploadView(APIView):
    """Upload files to a lesson plan."""
    permission_classes = [IsTeacherOrAbove]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, plan_id):
        try:
            plan = LessonPlan.objects.get(id=plan_id)
        except LessonPlan.DoesNotExist:
            return Response({"error": "Plan not found"}, status=404)

        files = request.FILES.getlist("files")
        if not files:
            return Response({"error": "No files provided"}, status=400)

        attachments = []
        for f in files:
            attachment = LessonPlanAttachment.objects.create(
                lesson_plan=plan,
                file=f,
                file_name=f.name,
                file_size=f.size,
                description=request.data.get("description", ""),
                uploaded_by=request.user,
            )
            attachments.append({
                "id": attachment.id,
                "file_name": attachment.file_name,
                "file_type": attachment.file_type,
                "file_size": attachment.file_size,
                "url": request.build_absolute_uri(attachment.file.url),
            })

        return Response({"uploaded": len(attachments), "attachments": attachments}, status=201)


class AttachmentListView(APIView):
    """List all attachments for a lesson plan."""
    permission_classes = [IsTeacherOrAbove]

    def get(self, request, plan_id):
        attachments = LessonPlanAttachment.objects.filter(lesson_plan_id=plan_id)
        data = LessonPlanAttachmentSerializer(
            attachments, many=True, context={"request": request}
        ).data
        return Response(data)


class AttachmentDeleteView(APIView):
    """Delete an attachment."""
    permission_classes = [IsTeacherOrAbove]

    def delete(self, request, pk):
        try:
            attachment = LessonPlanAttachment.objects.get(id=pk)
        except LessonPlanAttachment.DoesNotExist:
            return Response({"error": "Attachment not found"}, status=404)

        # Only the uploader or management can delete
        if attachment.uploaded_by != request.user and request.user.role not in ["hod", "deputy", "principal", "admin"]:
            return Response({"error": "Permission denied"}, status=403)

        attachment.file.delete()
        attachment.delete()
        return Response({"deleted": True})


# ─── IMPORT FROM EXCEL ───────────────────────────────────────────────────────

class ImportPlansFromExcelView(APIView):
    """
    Import lesson plans from an Excel/CSV file.
    Expected columns: day, period, title, objectives, activities, differentiation, resources
    """
    permission_classes = [IsTeacherOrAbove]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response({"error": "No file provided"}, status=400)

        ext = file.name.split(".")[-1].lower()

        try:
            if ext == "csv":
                rows = self._parse_csv(file)
            elif ext in ["xlsx", "xls"]:
                rows = self._parse_excel(file)
            else:
                return Response({"error": "Unsupported file format. Use .csv or .xlsx"}, status=400)
        except Exception as e:
            return Response({"error": f"Failed to parse file: {str(e)}"}, status=400)

        if not rows:
            return Response({"error": "No data rows found in file"}, status=400)

        from apps.schools.models import TimetableSlot, Timetable

        timetable = Timetable.objects.filter(
            school=request.user.school, status="active"
        ).first()

        if not timetable:
            return Response({"error": "No active timetable found"}, status=404)

        created = 0
        skipped = 0
        errors = []

        day_map = {
            "monday": "MON", "tuesday": "TUE", "wednesday": "WED",
            "thursday": "THU", "friday": "FRI",
            "mon": "MON", "tue": "TUE", "wed": "WED", "thu": "THU", "fri": "FRI",
        }

        for i, row in enumerate(rows, start=2):
            day_raw = str(row.get("day", "")).strip().lower()
            day = day_map.get(day_raw, day_raw.upper())

            try:
                period = int(row.get("period", 0))
            except (ValueError, TypeError):
                errors.append(f"Row {i}: Invalid period '{row.get('period')}'")
                skipped += 1
                continue

            title = str(row.get("title", "")).strip()
            objectives = str(row.get("objectives", "")).strip()
            activities = str(row.get("activities", "")).strip()

            if not title or not objectives:
                errors.append(f"Row {i}: Title and objectives are required")
                skipped += 1
                continue

            # Find the matching timetable slot
            slot = TimetableSlot.objects.filter(
                timetable=timetable,
                teacher=request.user,
                day=day,
                period=period,
            ).first()

            if not slot:
                errors.append(f"Row {i}: No slot found for {day} P{period}")
                skipped += 1
                continue

            # Check if plan already exists
            existing = LessonPlan.objects.filter(timetable_slot=slot).first()
            if existing:
                errors.append(f"Row {i}: Plan already exists for {day} P{period} — skipped")
                skipped += 1
                continue

            LessonPlan.objects.create(
                timetable_slot=slot,
                plan_type="official",
                status="draft",
                title=title,
                objectives=objectives,
                activities=activities,
                differentiation=str(row.get("differentiation", "")).strip(),
                resources_note=str(row.get("resources", "")).strip(),
                submitted_by=request.user,
            )
            created += 1

        return Response({
            "created": created,
            "skipped": skipped,
            "errors": errors,
            "total_rows": len(rows),
        })

    def _parse_csv(self, file):
        content = file.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(content))
        return [dict(row) for row in reader]

    def _parse_excel(self, file):
        try:
            import openpyxl
        except ImportError:
            raise Exception("openpyxl is required for Excel import. Install with: pip install openpyxl")

        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            return []

        # First row is headers
        headers = [str(h).strip().lower() if h else "" for h in rows[0]]
        data = []
        for row in rows[1:]:
            if not any(row):
                continue
            row_dict = {}
            for j, val in enumerate(row):
                if j < len(headers) and headers[j]:
                    row_dict[headers[j]] = val if val is not None else ""
            data.append(row_dict)
        return data


# ─── EXPORT PLANS ────────────────────────────────────────────────────────────

class ExportPlansView(APIView):
    """
    Export lesson plans to CSV or Excel.
    Exports based on the logged-in user's role and visible plans.
    """
    permission_classes = [IsTeacherOrAbove]

    def get(self, request):
        fmt = request.query_params.get("format", "csv")
        user = request.user

        # Get plans based on role
        qs = LessonPlan.objects.select_related(
            "timetable_slot", "timetable_slot__teacher",
            "timetable_slot__subject", "timetable_slot__classroom",
            "timetable_slot__classroom__grade",
            "submitted_by", "approved_by",
        ).order_by("timetable_slot__day", "timetable_slot__period")

        if user.role == "teacher":
            qs = qs.filter(timetable_slot__teacher=user)
        elif user.role == "hod":
            qs = qs.filter(timetable_slot__subject__hod=user)
        elif user.role == "grade_head":
            qs = qs.filter(timetable_slot__classroom__grade__grade_head=user)
        elif user.role in ["deputy", "principal", "admin"]:
            qs = qs.filter(timetable_slot__timetable__school=user.school)

        # Optional filters
        plan_status = request.query_params.get("status")
        if plan_status:
            qs = qs.filter(status=plan_status)

        plans = list(qs)

        if fmt == "xlsx":
            return self._export_xlsx(plans, user)
        else:
            return self._export_csv(plans, user)

    def _build_rows(self, plans):
        rows = []
        for plan in plans:
            slot = plan.timetable_slot
            rows.append({
                "Day": slot.get_day_display(),
                "Period": slot.period,
                "Time": f"{slot.start_time.strftime('%H:%M')}-{slot.end_time.strftime('%H:%M')}",
                "Subject": slot.subject.name,
                "Class": str(slot.classroom),
                "Teacher": slot.teacher.get_full_name() if slot.teacher else "",
                "Title": plan.title,
                "Status": plan.get_status_display(),
                "Objectives": plan.objectives,
                "Activities": plan.activities,
                "Differentiation": plan.differentiation,
                "Resources": plan.resources_note,
                "Version": plan.version,
                "Submitted By": plan.submitted_by.get_full_name() if plan.submitted_by else "",
                "Approved By": plan.approved_by.get_full_name() if plan.approved_by else "",
                "Created": plan.created_at.strftime("%Y-%m-%d %H:%M"),
            })
        return rows

    def _export_csv(self, plans, user):
        response = HttpResponse(content_type="text/csv")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        response["Content-Disposition"] = f'attachment; filename="lesson_plans_{timestamp}.csv"'

        rows = self._build_rows(plans)
        if not rows:
            return response

        writer = csv.DictWriter(response, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)
        return response

    def _export_xlsx(self, plans, user):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return Response(
                {"error": "openpyxl required. Install with: pip install openpyxl"},
                status=400,
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lesson Plans"

        rows = self._build_rows(plans)
        if not rows:
            ws.append(["No plans found"])
        else:
            # Header row
            headers = list(rows[0].keys())
            ws.append(headers)

            header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
            header_font = Font(color="FEF3C7", bold=True, size=11)
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Data rows
            for row in rows:
                ws.append(list(row.values()))

            # Auto-width columns
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        response["Content-Disposition"] = f'attachment; filename="lesson_plans_{timestamp}.xlsx"'
        wb.save(response)
        return response
